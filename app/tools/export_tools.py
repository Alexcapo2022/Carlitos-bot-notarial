import os
import uuid
import pandas as pd
from langchain.tools import tool
from sqlalchemy import text
from app.core.database import engine
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Directorio donde guardaremos los archivos generados
EXPORT_DIR = "/app/app/static/exports"
BASE_URL = "http://localhost:8001/static/exports"

@tool
def export_to_excel_tool(query: str) -> str:
    """
    Exporta los resultados de una consulta SQL a un archivo Excel (.xlsx).
    Uso: Pasa una consulta SQL válida (SELECT) como argumento. 
    Retorna la URL donde se puede descargar el archivo generado.
    NUNCA uses esto si el usuario no te ha pedido exportar, descargar o generar un Excel.
    """
    try:
        # Ejecutar la consulta y cargar resultados en un DataFrame de Pandas
        with engine.connect() as connection:
            df = pd.read_sql(text(query), connection)
        
        if df.empty:
            return "La consulta no devolvió resultados. No se generó el Excel."

        # Generar un nombre único para el archivo
        filename = f"reporte_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        # Exportar a Excel con diseño gerencial
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Reporte')
            workbook = writer.book
            worksheet = writer.sheets['Reporte']
            
            # Definir estilos
            header_fill = PatternFill(start_color="2A4B5C", end_color="2A4B5C", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            align_center = Alignment(horizontal="center", vertical="center")
            border_thin = Border(
                left=Side(style='thin', color='E2E8F0'), 
                right=Side(style='thin', color='E2E8F0'), 
                top=Side(style='thin', color='E2E8F0'), 
                bottom=Side(style='thin', color='E2E8F0')
            )
            
            # Aplicar estilo a la cabecera y ajustar ancho de columnas
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = border_thin
                
                # Calcular el ancho máximo basado en el contenido o el título
                max_len = len(str(column_title))
                for row_val in df.iloc[:, col_num-1]:
                    if len(str(row_val)) > max_len:
                        max_len = len(str(row_val))
                
                worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 60)
                
            # Aplicar bordes a las filas de datos
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = border_thin

        # Devolver la URL de descarga
        return f"¡Excel generado con éxito! El usuario puede descargarlo en: {BASE_URL}/{filename}"
        
    except Exception as e:
        return f"Error al generar el Excel: {str(e)}"
